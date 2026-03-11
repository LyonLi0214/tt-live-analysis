"""选择性 Excel 读取层 — 只读取目标活动相关的数据切片"""
import openpyxl
import pandas as pd
from datetime import datetime


def read_activity_data(file_path: str, activity_name: str) -> dict:
    """读取 Excel 模板，只提取与指定活动相关的数据。

    Returns:
        dict with keys: meta, daily, events_own, events_competitor, creators, historical, competitors
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    try:
        meta = _read_registry(wb, activity_name)
        daily = _read_daily_sheet(wb, activity_name)
        events_own, events_comp = _read_schedule(wb, activity_name)
        creators = _read_creators(wb, activity_name)
        historical = _read_historical(wb)
        competitors = _read_competitors(wb)
    finally:
        wb.close()

    return {
        "meta": meta,
        "daily": daily,
        "events_own": events_own,
        "events_competitor": events_comp,
        "creators": creators,
        "historical": historical,
        "competitors": competitors,
    }


def list_activities(file_path: str) -> list:
    """列出注册表中所有活动名称"""
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb["① 活动注册表"]
    activities = []
    for row in ws.iter_rows(min_row=5, max_col=1, values_only=True):
        val = row[0]
        if val and not str(val).startswith("——") and not str(val).startswith("↓"):
            activities.append(str(val))
    wb.close()
    return activities


def _read_registry(wb, activity_name: str) -> dict:
    """从①活动注册表读取目标活动的元信息"""
    ws = wb["① 活动注册表"]
    for row in ws.iter_rows(min_row=5, max_col=12, values_only=True):
        if row[0] and str(row[0]).strip() == activity_name:
            return {
                "name": str(row[0]).strip(),
                "start_date": _parse_date(row[1]),
                "end_date": _parse_date(row[2]),
                "days": _safe_int(row[3]),
                "type": str(row[4] or ""),
                "region": str(row[5] or ""),
                "budget": _safe_float(row[6]),
                "budget_traffic": _safe_float(row[7]),
                "budget_kol": _safe_float(row[8]),
                "budget_other": _safe_float(row[9]),
                "note": str(row[11] or ""),
            }
    raise ValueError(f"活动「{activity_name}」未在① 活动注册表中找到")


def _read_daily_sheet(wb, activity_name: str) -> pd.DataFrame:
    """读取以活动名命名的日粒度 Sheet"""
    if activity_name not in wb.sheetnames:
        raise ValueError(f"未找到日粒度数据 Sheet「{activity_name}」")

    ws = wb[activity_name]
    rows = []
    for row in ws.iter_rows(min_row=6, max_col=10, values_only=True):
        day_val = row[0]
        if day_val is None or not isinstance(day_val, (int, float)):
            continue
        rows.append({
            "day": int(day_val),
            "date": _parse_date(row[1]),
            "show_pv": _safe_float(row[2]),
            "show_uv": _safe_float(row[3]),
            "watch_pv": _safe_float(row[4]),
            "watch_uv": _safe_float(row[5]),
            "watch_duration": _safe_float(row[6]),
            "go_live_uv": _safe_float(row[7]),
            "live_duration": _safe_float(row[8]),
            "region": str(row[9] or ""),
        })

    if not rows:
        raise ValueError(f"Sheet「{activity_name}」中无有效数据行")

    return pd.DataFrame(rows)


def _read_schedule(wb, activity_name: str) -> tuple:
    """从③我方产品排期读取事件（分白区和蓝区）"""
    ws = wb["③ 我方产品排期"]
    events_own = []
    events_comp = []
    in_blue_area = False

    for row in ws.iter_rows(min_row=4, max_col=7, values_only=True):
        cell0 = str(row[0] or "")

        # 检测蓝色区域起始
        if "自动爬取" in cell0 or "竞品活动" in cell0:
            in_blue_area = True
            continue

        # 跳过标题行、分隔行、空行
        if not row[0] or cell0.startswith("——") or cell0.startswith("↓") or cell0 == "关联活动":
            continue
        if cell0.strip() == "关联活动\n【关联①】":
            continue

        # 筛选目标活动
        if cell0.strip() != activity_name:
            continue

        event = {
            "date": _parse_date(row[1]),
            "name": str(row[2] or ""),
            "type": str(row[3] or ""),
            "region": str(row[4] or ""),
            "impact": str(row[5] or ""),
            "note": str(row[6] or ""),
        }

        if in_blue_area:
            events_comp.append(event)
        else:
            events_own.append(event)

    return events_own, events_comp


def _read_creators(wb, activity_name: str) -> pd.DataFrame:
    """从④Top200创作者筛选目标活动的创作者"""
    ws = wb["④ Top200 创作者"]
    rows = []
    for row in ws.iter_rows(min_row=5, max_col=8, values_only=True):
        cell0 = str(row[0] or "")
        if cell0.startswith("——") or not row[0]:
            continue
        if cell0.strip() != activity_name:
            continue

        rows.append({
            "rank": _safe_int(row[1]),
            "link": str(row[2] or ""),
            "creator_id": str(row[3] or ""),
            "followers": str(row[4] or ""),
            "tier": str(row[5] or ""),
            "vertical": str(row[6] or ""),
            "note": str(row[7] or ""),
        })

    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _read_historical(wb) -> dict:
    """从⑤历史汇总读取所有活动汇总+基准线"""
    ws = wb["⑤ 历史汇总"]
    activities = []
    baselines = {}

    for row in ws.iter_rows(min_row=5, max_col=16, values_only=True):
        cell0 = str(row[0] or "")
        if not row[0] or cell0.startswith("▼") or cell0 == "基准线":
            continue
        if cell0.startswith("——"):
            continue

        # 基准线行
        if "P75" in cell0 or "P50" in cell0 or "P25" in cell0:
            key = cell0.split("(")[0].strip()
            baselines[key] = {
                "cpm": _safe_float(row[12]),
                "cpv": _safe_float(row[13]),
                "ctr": _safe_float(row[14]),
                "avg_watch_duration": _safe_float(row[15]),
            }
            continue

        activities.append({
            "name": cell0.strip(),
            "date_range": str(row[1] or ""),
            "days": _safe_int(row[2]),
            "region": str(row[3] or ""),
            "budget": _safe_float(row[4]),
            "show_pv": _safe_float(row[5]),
            "show_uv": _safe_float(row[6]),
            "watch_pv": _safe_float(row[7]),
            "watch_uv": _safe_float(row[8]),
            "watch_duration": _safe_float(row[9]),
            "go_live_uv": _safe_float(row[10]),
            "live_duration": _safe_float(row[11]),
        })

    return {"activities": activities, "baselines": baselines}


def _read_competitors(wb) -> list:
    """从⑥竞品监控读取竞品清单"""
    ws = wb["⑥ 竞品监控"]
    comps = []
    for row in ws.iter_rows(min_row=5, max_col=7, values_only=True):
        if not row[0] or str(row[0]).startswith("▼"):
            break
        comps.append({
            "game": str(row[0] or ""),
            "developer": str(row[1] or ""),
            "region": str(row[2] or ""),
            "channels": str(row[3] or ""),
            "reason": str(row[4] or ""),
            "priority": str(row[5] or ""),
        })
    return comps


def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m-%d", "%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _safe_int(val) -> int:
    if val is None:
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0
